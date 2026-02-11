import os
import math
from flask import Flask, render_template, jsonify, request
from openpyxl import load_workbook
from supabase import create_client, Client
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__, template_folder='../templates', static_folder='../static')

# ========== SUPABASE CONFIG ==========
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')
SUPABASE_BUCKET = os.getenv('SUPABASE_BUCKET', 'student-images')

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

STUDENTS_PER_PAGE = 15


# ========== READ EXCEL DATA ==========
def read_student_data():
    """Read students.xlsx and return list of student dicts"""
    xlsx_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'students.xlsx')

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    headers = []
    students = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            # Normalize headers
            headers = [str(h).strip() if h else '' for h in row]
            continue

        if not any(row):
            continue

        row_dict = {}
        for col_idx, cell_value in enumerate(row):
            if col_idx < len(headers):
                row_dict[headers[col_idx]] = cell_value

        # Map to our format — flexible key matching
        student = {
            'sno': find_value(row_dict, ['S.No', 'SNo', 'S.NO', 'Sno', 'sno']),
            'roll_number': find_value(row_dict, ['Roll Number', 'RollNumber', 'Roll No', 'Roll_Number', 'roll_number']),
            'full_name': find_value(row_dict, ['Full Name', 'FullName', 'Name', 'Full_Name', 'full_name']),
            'branch': find_value(row_dict, ['Branch', 'branch', 'BRANCH']),
            'rating': find_value(row_dict, ['Rating (1-5)', 'Rating', 'rating', 'Rating(1-5)']),
            'comment': find_value(row_dict, ['Review/Comment', 'Comment', 'Review', 'comment', 'review']),
            'image_path': find_value(row_dict, ['Image Path', 'ImagePath', 'Image_Path', 'image_path', 'Photo']),
        }

        # Clean up rating
        try:
            student['rating'] = float(student['rating']) if student['rating'] else 0
        except (ValueError, TypeError):
            student['rating'] = 0

        # Clean up strings
        student['roll_number'] = str(student['roll_number']).strip() if student['roll_number'] else ''
        student['full_name'] = str(student['full_name']).strip() if student['full_name'] else ''
        student['branch'] = str(student['branch']).strip() if student['branch'] else ''
        student['comment'] = str(student['comment']).strip() if student['comment'] else ''
        student['image_path'] = str(student['image_path']).strip() if student['image_path'] else ''

        # Generate Supabase public URL for image
        if student['image_path']:
            # Extract filename from path like "student_images/25B21A4218.jpg"
            filename = student['image_path'].split('/')[-1]
            student['image_url'] = get_supabase_image_url(filename)
        else:
            student['image_url'] = ''

        students.append(student)

    wb.close()
    return students


def find_value(row_dict, possible_keys):
    """Find value by trying multiple possible column names"""
    for key in possible_keys:
        if key in row_dict:
            return row_dict[key]
    return None


def get_supabase_image_url(filename):
    """Generate public URL for image stored in Supabase Storage"""
    return f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{filename}"


# ========== ROUTES ==========

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')


@app.route('/api/students')
def get_students():
    """API: Get paginated students with optional search"""
    try:
        page = int(request.args.get('page', 1))
        search = request.args.get('search', '').strip().lower()

        students = read_student_data()

        # Filter by search query
        if search:
            students = [
                s for s in students
                if search in s['roll_number'].lower()
                or search in s['full_name'].lower()
                or search in s['branch'].lower()
            ]

        total_students = len(students)
        total_pages = max(1, math.ceil(total_students / STUDENTS_PER_PAGE))
        page = max(1, min(page, total_pages))

        start = (page - 1) * STUDENTS_PER_PAGE
        end = start + STUDENTS_PER_PAGE
        paginated = students[start:end]

        return jsonify({
            'success': True,
            'students': paginated,
            'current_page': page,
            'total_pages': total_pages,
            'total_students': total_students,
            'per_page': STUDENTS_PER_PAGE
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/student/<roll_number>')
def get_student(roll_number):
    """API: Get single student by roll number"""
    try:
        students = read_student_data()
        student = next((s for s in students if s['roll_number'] == roll_number), None)

        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        return jsonify({'success': True, 'student': student})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ========== UPLOAD HELPER (Run once to upload images to Supabase) ==========
@app.route('/api/upload-images', methods=['POST'])
def upload_images():
    """Upload all images from local folder to Supabase Storage"""
    try:
        images_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'student_images'
        )

        if not os.path.exists(images_dir):
            return jsonify({'success': False, 'message': 'student_images folder not found'})

        uploaded = []
        errors = []

        for filename in os.listdir(images_dir):
            filepath = os.path.join(images_dir, filename)
            if os.path.isfile(filepath):
                try:
                    with open(filepath, 'rb') as f:
                        supabase.storage.from_(SUPABASE_BUCKET).upload(
                            filename,
                            f.read(),
                            {"content-type": "image/jpeg"}
                        )
                    uploaded.append(filename)
                except Exception as e:
                    errors.append({'file': filename, 'error': str(e)})

        return jsonify({
            'success': True,
            'uploaded': len(uploaded),
            'errors': errors
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)
